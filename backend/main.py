import hashlib
import io
import os
import random
import shutil
import uuid
from datetime import datetime
from typing import List, Optional

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from database import engine, get_db, Base
from models import User, LanguagePair, Sentence, TestSession, SentenceRating
from auth import (
    hash_password, verify_password, create_token,
    get_current_user, get_admin_user,
)

Base.metadata.create_all(bind=engine)

app = FastAPI(title="Timekettle Translation Evaluation System")
_cors_origins_env = os.getenv("CORS_ORIGINS", "").strip()
_cors_origins = (
    [o.strip() for o in _cors_origins_env.split(",") if o.strip()]
    if _cors_origins_env
    else ["*"]
)
# NOTE: Browsers disallow `Access-Control-Allow-Origin: *` when credentials are enabled.
# This app uses Bearer tokens, so credentials are not required by default.
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=False if _cors_origins == ["*"] else True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


def _ensure_audio_folders(pair_code: str) -> None:
    """确保语言对的音频文件夹结构存在"""
    base = os.path.join(UPLOAD_DIR, pair_code)
    for folder in ["source", "engine1", "engine2"]:
        os.makedirs(os.path.join(base, folder), exist_ok=True)


def _init_admin(db: Session):
    admin_user_id = os.getenv("ADMIN_USER_ID", "admin").strip() or "admin"
    admin_password = os.getenv("ADMIN_PASSWORD", "admin123")

    admin = db.query(User).filter(User.user_id == admin_user_id).first()
    if not admin:
        admin = User(
            user_id=admin_user_id,
            password_hash=hash_password(admin_password),
            is_admin=True,
        )
        db.add(admin)
        db.commit()


@app.on_event("startup")
def startup():
    db = next(get_db())
    _init_admin(db)
    db.close()


# ─── Schemas ────────────────────────────────────────────

class LoginReq(BaseModel):
    user_id: str
    password: str

class LoginResp(BaseModel):
    token: str
    is_admin: bool
    user_id: str

class LangPairCreate(BaseModel):
    source_lang: str
    target_lang: str
    pair_code: str
    display_name: str

class TesterCreate(BaseModel):
    user_id: str
    password: str
    language_pair_ids: Optional[List[int]] = None  # 可测语言对 ID 列表，空或不传表示不限制（全部可测）

class RatingSubmit(BaseModel):
    sentence_id: int
    sentence_index: int
    user_rating: int          # -2 ~ +2
    left_play_count: int
    right_play_count: int
    duration_seconds: float

class ProgressUpdate(BaseModel):
    current_index: int


# ─── Auth ───────────────────────────────────────────────

@app.post("/api/login", response_model=LoginResp)
def login(req: LoginReq, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.user_id == req.user_id).first()
    if not user or not verify_password(req.password, user.password_hash):
        raise HTTPException(400, "账号或密码错误")
    token = create_token(user.user_id, user.is_admin)
    return LoginResp(token=token, is_admin=user.is_admin, user_id=user.user_id)


# ─── Tester: language pair listing ──────────────────────

@app.get("/api/language-pairs")
def list_language_pairs(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    pairs = db.query(LanguagePair).all()
    # 测试人员若配置了可测语言对，则只返回这些
    if not user.is_admin:
        db.refresh(user)  # 加载 allowed_language_pairs
        if user.allowed_language_pairs:
            allowed_ids = {lp.id for lp in user.allowed_language_pairs}
            pairs = [p for p in pairs if p.id in allowed_ids]
    result = []
    for lp in pairs:
        sentence_count = db.query(func.count(Sentence.id)).filter(Sentence.language_pair_id == lp.id).scalar()
        session = (
            db.query(TestSession)
            .filter(TestSession.user_id == user.id, TestSession.language_pair_id == lp.id, TestSession.status != "voided")
            .first()
        )
        status = "not_started"
        if session:
            status = session.status
        result.append({
            "id": lp.id,
            "pair_code": lp.pair_code,
            "display_name": lp.display_name,
            "source_lang": lp.source_lang,
            "target_lang": lp.target_lang,
            "sentence_count": sentence_count,
            "status": status,
        })
    return result


# ─── Tester: start / resume test ────────────────────────

def _engine_side(seed: int, sentence_index: int):
    """Deterministic left/right assignment based on seed + index."""
    rng = random.Random(seed ^ sentence_index)
    if rng.random() < 0.5:
        return "self_research", "iflytek"
    return "iflytek", "self_research"


@app.post("/api/test/start")
def start_test(
    language_pair_id: int = Query(...),
    restart: bool = Query(False),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    lp = db.query(LanguagePair).filter(LanguagePair.id == language_pair_id).first()
    if not lp:
        raise HTTPException(404, "语言对不存在")
    # 测试人员只能测试已分配的语言对
    if not user.is_admin:
        db.refresh(user)
        if user.allowed_language_pairs:
            allowed_ids = {p.id for p in user.allowed_language_pairs}
            if language_pair_id not in allowed_ids:
                raise HTTPException(403, "您没有该语言对的测试权限")

    existing = (
        db.query(TestSession)
        .filter(TestSession.user_id == user.id, TestSession.language_pair_id == language_pair_id, TestSession.status != "voided")
        .first()
    )

    if existing and existing.status == "completed":
        raise HTTPException(400, "该语言对已完成测试，不可重复测试")

    if existing and not restart:
        return {
            "session_id": existing.id,
            "current_index": existing.current_index,
            "status": existing.status,
            "resumed": True,
        }

    if existing and restart:
        existing.status = "voided"
        db.commit()

    seed = random.randint(0, 2**31)
    session = TestSession(
        user_id=user.id,
        language_pair_id=language_pair_id,
        randomization_seed=seed,
    )
    db.add(session)
    db.commit()
    db.refresh(session)
    return {
        "session_id": session.id,
        "current_index": 1,
        "status": "in_progress",
        "resumed": False,
    }


@app.get("/api/test/sentence")
def get_sentence(
    session_id: int = Query(...),
    index: int = Query(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    session = db.query(TestSession).filter(TestSession.id == session_id, TestSession.user_id == user.id).first()
    if not session:
        raise HTTPException(404, "测试会话不存在")

    sentence = (
        db.query(Sentence)
        .filter(Sentence.language_pair_id == session.language_pair_id, Sentence.sentence_index == index)
        .first()
    )
    if not sentence:
        raise HTTPException(404, "句子不存在")

    total = db.query(func.count(Sentence.id)).filter(Sentence.language_pair_id == session.language_pair_id).scalar()
    engine_left, engine_right = _engine_side(session.randomization_seed, index)

    existing_rating = (
        db.query(SentenceRating)
        .filter(SentenceRating.test_session_id == session_id, SentenceRating.sentence_index == index)
        .first()
    )

    def _audio_url(path):
        if not path:
            return None
        return f"/api/audio/{path}"

    if engine_left == "self_research":
        left_trans = sentence.engine1_translation_text
        left_recog = sentence.engine1_recognition_text
        left_audio = _audio_url(sentence.engine1_audio_path)
        right_trans = sentence.engine2_translation_text
        right_recog = sentence.engine2_recognition_text
        right_audio = _audio_url(sentence.engine2_audio_path)
    else:
        left_trans = sentence.engine2_translation_text
        left_recog = sentence.engine2_recognition_text
        left_audio = _audio_url(sentence.engine2_audio_path)
        right_trans = sentence.engine1_translation_text
        right_recog = sentence.engine1_recognition_text
        right_audio = _audio_url(sentence.engine1_audio_path)

    return {
        "sentence_id": sentence.id,
        "sid": sentence.sid,
        "sentence_index": index,
        "total": total,
        "source_text": sentence.source_text,
        "source_audio": _audio_url(sentence.source_audio_path),
        "source_audio_duration": sentence.source_audio_duration,
        "left_translation_text": left_trans,
        "left_recognition_text": left_recog,
        "left_audio": left_audio,
        "right_translation_text": right_trans,
        "right_recognition_text": right_recog,
        "right_audio": right_audio,
        "engine_left": engine_left,
        "engine_right": engine_right,
        "existing_rating": existing_rating.user_rating if existing_rating else None,
        "existing_left_play_count": existing_rating.left_play_count if existing_rating else 0,
        "existing_right_play_count": existing_rating.right_play_count if existing_rating else 0,
    }


@app.post("/api/test/rate")
def submit_rating(
    session_id: int = Query(...),
    body: RatingSubmit = ...,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    session = db.query(TestSession).filter(TestSession.id == session_id, TestSession.user_id == user.id).first()
    if not session:
        raise HTTPException(404, "测试会话不存在")
    if session.status == "completed":
        raise HTTPException(400, "测试已完成，不可修改")

    engine_left, engine_right = _engine_side(session.randomization_seed, body.sentence_index)

    existing = (
        db.query(SentenceRating)
        .filter(SentenceRating.test_session_id == session_id, SentenceRating.sentence_index == body.sentence_index)
        .first()
    )
    if existing:
        existing.user_rating = body.user_rating
        existing.left_play_count = body.left_play_count
        existing.right_play_count = body.right_play_count
        existing.duration_seconds = body.duration_seconds
        existing.rated_at = datetime.utcnow()
    else:
        rating = SentenceRating(
            test_session_id=session_id,
            sentence_id=body.sentence_id,
            sentence_index=body.sentence_index,
            engine_left=engine_left,
            engine_right=engine_right,
            left_play_count=body.left_play_count,
            right_play_count=body.right_play_count,
            user_rating=body.user_rating,
            duration_seconds=body.duration_seconds,
            rated_at=datetime.utcnow(),
        )
        db.add(rating)

    session.current_index = body.sentence_index + 1
    db.commit()
    return {"ok": True}


@app.post("/api/test/pause")
def pause_test(
    session_id: int = Query(...),
    body: ProgressUpdate = ...,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    session = db.query(TestSession).filter(TestSession.id == session_id, TestSession.user_id == user.id).first()
    if not session:
        raise HTTPException(404, "测试会话不存在")
    session.current_index = body.current_index
    db.commit()
    return {"ok": True}


@app.post("/api/test/complete")
def complete_test(
    session_id: int = Query(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    session = db.query(TestSession).filter(TestSession.id == session_id, TestSession.user_id == user.id).first()
    if not session:
        raise HTTPException(404, "测试会话不存在")
    session.status = "completed"
    session.completed_at = datetime.utcnow()
    db.commit()

    ratings = db.query(SentenceRating).filter(SentenceRating.test_session_id == session_id).all()
    total_score = sum(r.user_rating or 0 for r in ratings)
    lp = db.query(LanguagePair).filter(LanguagePair.id == session.language_pair_id).first()
    return {
        "ok": True,
        "total_score": total_score,
        "count": len(ratings),
        "language_pair": lp.display_name if lp else "",
    }


# ─── Audio serving ──────────────────────────────────────

_AUDIO_MIME = {
    ".mp3": "audio/mpeg",
    ".wav": "audio/wav",
    ".m4a": "audio/mp4",
    ".aac": "audio/aac",
    ".ogg": "audio/ogg",
    ".webm": "audio/webm",
}

@app.get("/api/audio/{path:path}")
def serve_audio(path: str):
    full = os.path.join(UPLOAD_DIR, path)
    if not os.path.isfile(full):
        raise HTTPException(404, "Audio not found")
    ext = os.path.splitext(path)[1].lower()
    media_type = _AUDIO_MIME.get(ext, "audio/mpeg")
    return FileResponse(full, media_type=media_type)


# ─── Admin: language pair CRUD ──────────────────────────

@app.get("/api/admin/language-pairs")
def admin_list_pairs(admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    pairs = db.query(LanguagePair).all()
    result = []
    for lp in pairs:
        cnt = db.query(func.count(Sentence.id)).filter(Sentence.language_pair_id == lp.id).scalar()
        result.append({
            "id": lp.id, "pair_code": lp.pair_code, "display_name": lp.display_name,
            "source_lang": lp.source_lang, "target_lang": lp.target_lang,
            "sentence_count": cnt,
        })
    return result


@app.post("/api/admin/language-pairs")
def admin_create_pair(body: LangPairCreate, admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    existing = db.query(LanguagePair).filter(LanguagePair.pair_code == body.pair_code).first()
    if existing:
        raise HTTPException(400, "语言对编码已存在")
    lp = LanguagePair(**body.model_dump())
    db.add(lp)
    db.commit()
    db.refresh(lp)
    return {"id": lp.id, "pair_code": lp.pair_code}


@app.delete("/api/admin/language-pairs/{pair_id}")
def admin_delete_pair(pair_id: int, admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    lp = db.query(LanguagePair).filter(LanguagePair.id == pair_id).first()
    if not lp:
        raise HTTPException(404, "Not found")
    db.delete(lp)
    db.commit()
    return {"ok": True}


# 语言对 Excel 模板：表头与新增表单一致
PAIR_TEMPLATE_HEADERS = ["源语言", "目标语言", "编码", "显示名称"]

# 魔法导入 Excel 模板
MAGIC_IMPORT_HEADERS = [
    '语料ID', '原始文本', '音频时长(秒)', '源语言', '目标语言',
    '引擎1翻译文本', '引擎1识别文本', '引擎2翻译文本', '引擎2识别文本',
    '编码', '显示名称'
]


@app.get("/api/admin/language-pairs/template")
def admin_download_pair_template(admin: User = Depends(get_admin_user)):
    """下载语言对导入模板（Excel），表头为：源语言、目标语言、编码、显示名称。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "语言对"
    ws.append(PAIR_TEMPLATE_HEADERS)
    ws.append(["zh", "en", "zh-en", "中文 → 英语"])  # 示例行
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=language_pairs_template.xlsx"},
    )


@app.get("/api/admin/magic-import-template")
def admin_download_magic_import_template(admin: User = Depends(get_admin_user)):
    """下载魔法导入模板（Excel）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "魔法导入"
    ws.append(MAGIC_IMPORT_HEADERS)
    # 示例行
    ws.append([
        "001",  # 语料ID
        "Hello world",  # 原始文本
        3.5,  # 音频时长
        "en",  # 源语言
        "zh",  # 目标语言
        "你好世界",  # 引擎1翻译文本
        "",  # 引擎1识别文本
        "",  # 引擎2翻译文本
        "",  # 引擎2识别文本
        "en-zh",  # 编码
        "英语 → 中文"  # 显示名称
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=magic_import_template.xlsx"},
    )


@app.post("/api/admin/language-pairs/import-excel")
async def admin_import_pairs_excel(
    file: UploadFile = File(...),
    admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """根据 Excel 批量导入语言对。表头：源语言、目标语言、编码、显示名称。"""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "请上传 .xlsx 或 .xls 文件")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "skipped": 0, "errors": [{"row": 1, "message": "文件无数据"}]}
    # 首行为表头，定位列索引
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_map = {}
    for i, name in enumerate(header):
        if name in PAIR_TEMPLATE_HEADERS:
            col_map[name] = i
    if len(col_map) != 4:
        return {
            "imported": 0,
            "skipped": 0,
            "errors": [{"row": 1, "message": f"表头需包含：{PAIR_TEMPLATE_HEADERS}"}],
        }
    imported = 0
    skipped = 0
    errors = []
    for idx, row in enumerate(rows[1:], start=2):
        if not row:
            continue
        source_lang = str(row[col_map["源语言"]] or "").strip()
        target_lang = str(row[col_map["目标语言"]] or "").strip()
        pair_code = str(row[col_map["编码"]] or "").strip()
        display_name = str(row[col_map["显示名称"]] or "").strip()
        if not pair_code or not display_name:
            if pair_code or display_name or any(str(x).strip() for x in row if x is not None):
                errors.append({"row": idx, "message": "编码和显示名称不能为空"})
            continue
        existing = db.query(LanguagePair).filter(LanguagePair.pair_code == pair_code).first()
        if existing:
            skipped += 1
            continue
        lp = LanguagePair(
            source_lang=source_lang or "",
            target_lang=target_lang or "",
            pair_code=pair_code,
            display_name=display_name,
        )
        db.add(lp)
        imported += 1
    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors}


# ─── Admin: sentence CRUD ──────────────────────────────

def _save_upload(file: UploadFile, pair_code: str) -> str:
    subdir = os.path.join(UPLOAD_DIR, pair_code)
    os.makedirs(subdir, exist_ok=True)
    ext = os.path.splitext(file.filename or "audio.wav")[1]
    name = f"{uuid.uuid4().hex}{ext}"
    dest = os.path.join(subdir, name)
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)
    return f"{pair_code}/{name}"


@app.get("/api/admin/sentences")
def admin_list_sentences(
    language_pair_id: int = Query(...),
    admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    sentences = (
        db.query(Sentence)
        .filter(Sentence.language_pair_id == language_pair_id)
        .order_by(Sentence.sentence_index)
        .all()
    )
    return [
        {
            "id": s.id,
            "sid": s.sid,
            "sentence_index": s.sentence_index,
            "source_text": s.source_text,
            "source_audio_path": s.source_audio_path,
            "source_audio_duration": s.source_audio_duration,
            "engine1_translation_text": s.engine1_translation_text,
            "engine1_recognition_text": s.engine1_recognition_text,
            "engine1_audio_path": s.engine1_audio_path,
            "engine2_translation_text": s.engine2_translation_text,
            "engine2_recognition_text": s.engine2_recognition_text,
            "engine2_audio_path": s.engine2_audio_path,
        }
        for s in sentences
    ]


@app.post("/api/admin/sentences")
async def admin_create_sentence(
    language_pair_id: int = Form(...),
    sid: str = Form(...),
    source_text: str = Form(...),
    source_audio_duration: float = Form(0),
    engine1_translation_text: str = Form(""),
    engine1_recognition_text: str = Form(""),
    engine2_translation_text: str = Form(""),
    engine2_recognition_text: str = Form(""),
    source_audio: Optional[UploadFile] = File(None),
    engine1_audio: Optional[UploadFile] = File(None),
    engine2_audio: Optional[UploadFile] = File(None),
    admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    lp = db.query(LanguagePair).filter(LanguagePair.id == language_pair_id).first()
    if not lp:
        raise HTTPException(404, "语言对不存在")

    existing = db.query(Sentence).filter(
        Sentence.language_pair_id == language_pair_id, Sentence.sid == sid
    ).first()
    if existing:
        raise HTTPException(400, f"语料ID '{sid}' 已存在，请勿重复添加")

    max_idx = db.query(func.max(Sentence.sentence_index)).filter(
        Sentence.language_pair_id == language_pair_id
    ).scalar() or 0

    s = Sentence(
        language_pair_id=language_pair_id,
        sentence_index=max_idx + 1,
        sid=sid,
        source_text=source_text,
        source_audio_duration=source_audio_duration,
        engine1_translation_text=engine1_translation_text,
        engine1_recognition_text=engine1_recognition_text,
        engine2_translation_text=engine2_translation_text,
        engine2_recognition_text=engine2_recognition_text,
    )
    if source_audio:
        s.source_audio_path = _save_upload(source_audio, lp.pair_code)
    if engine1_audio:
        s.engine1_audio_path = _save_upload(engine1_audio, lp.pair_code)
    if engine2_audio:
        s.engine2_audio_path = _save_upload(engine2_audio, lp.pair_code)

    db.add(s)
    db.commit()
    db.refresh(s)
    return {"id": s.id}


@app.delete("/api/admin/sentences/{sentence_id}")
def admin_delete_sentence(sentence_id: int, admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    s = db.query(Sentence).filter(Sentence.id == sentence_id).first()
    if not s:
        raise HTTPException(404)
    db.delete(s)
    db.commit()
    return {"ok": True}


class BatchDeleteSentencesReq(BaseModel):
    ids: List[int]


@app.post("/api/admin/sentences/batch-delete")
def admin_batch_delete_sentences(
    body: BatchDeleteSentencesReq,
    admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    if not body.ids:
        return {"deleted": 0}
    deleted = db.query(Sentence).filter(Sentence.id.in_(body.ids)).delete(synchronize_session=False)
    db.commit()
    return {"deleted": deleted}


# ─── Admin: Excel batch import ──────────────────────────

@app.post("/api/admin/sentences/import-excel")
async def admin_import_excel(
    language_pair_id: int = Form(...),
    file: UploadFile = File(...),
    admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    lp = db.query(LanguagePair).filter(LanguagePair.id == language_pair_id).first()
    if not lp:
        raise HTTPException(404, "语言对不存在")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    max_idx = db.query(func.max(Sentence.sentence_index)).filter(
        Sentence.language_pair_id == language_pair_id
    ).scalar() or 0

    count = 0
    skipped = 0
    filtered = 0
    for row in rows:
        if not row or not row[0]:
            continue
        row_sid = str(row[0]).strip()
        # 新格式：col3=源语言, col4=目标语言，不匹配当前语言对则跳过
        row_source_lang = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        row_target_lang = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
        if row_source_lang and row_target_lang:
            if row_source_lang != lp.source_lang or row_target_lang != lp.target_lang:
                filtered += 1
                continue
        existing = db.query(Sentence).filter(
            Sentence.language_pair_id == language_pair_id, Sentence.sid == row_sid
        ).first()
        if existing:
            skipped += 1
            continue
        max_idx += 1
        s = Sentence(
            language_pair_id=language_pair_id,
            sentence_index=max_idx,
            sid=row_sid,
            source_text=str(row[1] or ""),
            source_audio_duration=float(row[2] or 0),
            engine1_translation_text=str(row[5] or "") if len(row) > 5 else "",
            engine1_recognition_text=str(row[6] or "") if len(row) > 6 else "",
            engine2_translation_text=str(row[7] or "") if len(row) > 7 else "",
            engine2_recognition_text=str(row[8] or "") if len(row) > 8 else "",
        )
        db.add(s)
        count += 1
    db.commit()
    return {"imported": count, "skipped": skipped, "filtered": filtered}


# ─── Admin: scan audio files ────────────────────────────

@app.post("/api/admin/sentences/scan-audio")
def admin_scan_audio(
    language_pair_id: int = Query(...),
    admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    import glob as globmod

    lp = db.query(LanguagePair).filter(LanguagePair.id == language_pair_id).first()
    if not lp:
        raise HTTPException(404, "语言对不存在")

    base = os.path.join(UPLOAD_DIR, lp.pair_code)
    folders = {
        "source": "source_audio_path",
        "engine1": "engine1_audio_path",
        "engine2": "engine2_audio_path",
    }

    matched = 0
    not_found = 0

    for folder_name, db_field in folders.items():
        folder_path = os.path.join(base, folder_name)
        if not os.path.isdir(folder_path):
            continue

        for filepath in globmod.glob(os.path.join(folder_path, "*")):
            filename = os.path.basename(filepath)
            name_part = os.path.splitext(filename)[0]
            if not name_part:
                continue

            # 优先精确匹配，再尝试去掉前导零匹配（如文件 001 → sid 1）
            sentence = (
                db.query(Sentence)
                .filter(Sentence.language_pair_id == language_pair_id, Sentence.sid == name_part)
                .first()
            )
            if not sentence and name_part.isdigit():
                stripped = str(int(name_part))
                sentence = (
                    db.query(Sentence)
                    .filter(Sentence.language_pair_id == language_pair_id, Sentence.sid == stripped)
                    .first()
                )
            if sentence:
                rel_path = f"{lp.pair_code}/{folder_name}/{filename}"
                setattr(sentence, db_field, rel_path)
                matched += 1
            else:
                not_found += 1

    db.commit()
    return {"matched": matched, "not_found": not_found}


# ─── Admin: 魔法导入 ───────────────────────────────────

@app.post("/api/admin/magic-import-excel")
async def admin_magic_import_excel(
    file: UploadFile = File(...),
    admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """魔法导入：一次上传同时创建语言对和导入语料"""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "请上传 .xlsx 或 .xls 文件")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # 按语言对分组数据
    pairs_data = {}  # {pair_code: {lp: LanguagePair, sentences: []}}
    stats = {"pairs_created": 0, "pairs_existed": 0, "sentences_imported": 0, "sentences_skipped": 0}

    for row in rows:
        if not row or not row[0]:
            continue
        # 解析行数据
        row_sid = str(row[0]).strip()
        # 去掉 "/" 之前的内容和扩展名
        if '/' in row_sid:
            row_sid = row_sid.split('/')[-1]
        if '.' in row_sid:
            row_sid = row_sid.rsplit('.', 1)[0]

        source_text = str(row[1] or "").strip()
        source_audio_duration = float(row[2] or 0)
        source_lang = str(row[3]).strip() if row[3] else ""
        target_lang = str(row[4]).strip() if row[4] else ""
        engine1_translation = str(row[5] or "") if len(row) > 5 else ""
        engine1_recognition = str(row[6] or "") if len(row) > 6 else ""
        engine2_translation = str(row[7] or "") if len(row) > 7 else ""
        engine2_recognition = str(row[8] or "") if len(row) > 8 else ""
        pair_code = str(row[9] or "").strip() if len(row) > 9 else ""
        display_name = str(row[10] or "").strip() if len(row) > 10 else ""

        if not pair_code:
            continue

        # 获取或创建语言对
        if pair_code not in pairs_data:
            lp = db.query(LanguagePair).filter(LanguagePair.pair_code == pair_code).first()
            if not lp:
                lp = LanguagePair(
                    source_lang=source_lang,
                    target_lang=target_lang,
                    pair_code=pair_code,
                    display_name=display_name or f"{source_lang} → {target_lang}",
                )
                db.add(lp)
                db.flush()
                stats["pairs_created"] += 1
            else:
                stats["pairs_existed"] += 1
            # 无论语言对新建还是已存在，都确保音频文件夹存在
            _ensure_audio_folders(pair_code)
            # 获取当前最大 index 作为起始点
            max_idx = db.query(func.max(Sentence.sentence_index)).filter(
                Sentence.language_pair_id == lp.id
            ).scalar() or 0
            pairs_data[pair_code] = {"lp": lp, "next_index": max_idx + 1}

        # 检查语料是否已存在
        existing = db.query(Sentence).filter(
            Sentence.language_pair_id == pairs_data[pair_code]["lp"].id,
            Sentence.sid == row_sid
        ).first()
        if existing:
            stats["sentences_skipped"] += 1
            continue

        # 使用内存中的计数器
        next_idx = pairs_data[pair_code]["next_index"]

        s = Sentence(
            language_pair_id=pairs_data[pair_code]["lp"].id,
            sentence_index=next_idx,
            sid=row_sid,
            source_text=source_text,
            source_audio_duration=source_audio_duration,
            engine1_translation_text=engine1_translation,
            engine1_recognition_text=engine1_recognition,
            engine2_translation_text=engine2_translation,
            engine2_recognition_text=engine2_recognition,
        )
        db.add(s)
        pairs_data[pair_code]["next_index"] = next_idx + 1
        stats["sentences_imported"] += 1

    db.commit()
    return stats


# ─── Admin: 扫描所有语言对音频 ───────────────────────────

@app.post("/api/admin/scan-all-audio")
def admin_scan_all_audio(
    admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """扫描所有语言对的音频文件，匹配语料ID"""
    import glob as globmod

    all_pairs = db.query(LanguagePair).all()
    total_matched = 0
    total_not_found = 0
    results = []

    for lp in all_pairs:
        base = os.path.join(UPLOAD_DIR, lp.pair_code)
        if not os.path.isdir(base):
            results.append({
                "pair_code": lp.pair_code,
                "matched": 0,
                "not_found": 0,
                "status": "folder_not_found"
            })
            continue

        folders = {
            "source": "source_audio_path",
            "engine1": "engine1_audio_path",
            "engine2": "engine2_audio_path",
        }

        matched = 0
        not_found = 0

        for folder_name, db_field in folders.items():
            folder_path = os.path.join(base, folder_name)
            if not os.path.isdir(folder_path):
                continue

            for filepath in globmod.glob(os.path.join(folder_path, "*")):
                filename = os.path.basename(filepath)
                name_part = os.path.splitext(filename)[0]
                if not name_part:
                    continue

                # 优先精确匹配，再尝试去掉前导零匹配
                sentence = (
                    db.query(Sentence)
                    .filter(Sentence.language_pair_id == lp.id, Sentence.sid == name_part)
                    .first()
                )
                if not sentence and name_part.isdigit():
                    stripped = str(int(name_part))
                    sentence = (
                        db.query(Sentence)
                        .filter(Sentence.language_pair_id == lp.id, Sentence.sid == stripped)
                        .first()
                    )
                if sentence:
                    rel_path = f"{lp.pair_code}/{folder_name}/{filename}"
                    setattr(sentence, db_field, rel_path)
                    matched += 1
                else:
                    not_found += 1

        if matched > 0 or not_found > 0:
            db.commit()

        total_matched += matched
        total_not_found += not_found
        results.append({
            "pair_code": lp.pair_code,
            "matched": matched,
            "not_found": not_found,
            "status": "ok"
        })

    return {
        "total_matched": total_matched,
        "total_not_found": total_not_found,
        "pairs": results
    }


# ─── Admin: tester CRUD ────────────────────────────────

@app.get("/api/admin/testers")
def admin_list_testers(admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    testers = db.query(User).filter(User.is_admin == False).all()
    result = []
    for t in testers:
        db.refresh(t)  # 加载 allowed_language_pairs
        sessions = db.query(TestSession).filter(TestSession.user_id == t.id, TestSession.status != "voided").all()
        completed_pairs = [s for s in sessions if s.status == "completed"]
        lp_codes = []
        for s in sessions:
            lp = db.query(LanguagePair).filter(LanguagePair.id == s.language_pair_id).first()
            if lp and s.status == "completed":
                lp_codes.append(lp.pair_code)

        if completed_pairs:
            status = "已完成" if len(completed_pairs) == len(sessions) and sessions else "测试中"
        elif sessions:
            status = "测试中"
        else:
            status = "未开始"

        allowed_pairs = [{"id": lp.id, "display_name": lp.display_name} for lp in t.allowed_language_pairs]

        result.append({
            "id": t.id,
            "user_id": t.user_id,
            "status": status,
            "completed_pairs": lp_codes,
            "allowed_language_pairs": allowed_pairs,
            "created_at": t.created_at.isoformat() if t.created_at else None,
        })
    return result


@app.post("/api/admin/testers")
def admin_create_tester(body: TesterCreate, admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    existing = db.query(User).filter(User.user_id == body.user_id).first()
    if existing:
        raise HTTPException(400, "账号已存在")
    user = User(user_id=body.user_id, password_hash=hash_password(body.password), is_admin=False)
    db.add(user)
    db.commit()
    db.refresh(user)
    if body.language_pair_ids:
        pairs = db.query(LanguagePair).filter(LanguagePair.id.in_(body.language_pair_ids)).all()
        user.allowed_language_pairs = pairs
        db.commit()
    return {"id": user.id, "user_id": user.user_id}


@app.post("/api/admin/testers/import-excel")
async def admin_import_testers_excel(
    file: UploadFile = File(...),
    admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    count = 0
    skipped = 0
    for row in rows:
        if not row or not row[0]:
            continue
        uid = str(row[0]).strip()
        pwd = str(row[1]).strip() if len(row) > 1 and row[1] else "123456"
        existing = db.query(User).filter(User.user_id == uid).first()
        if existing:
            skipped += 1
            continue
        user = User(user_id=uid, password_hash=hash_password(pwd), is_admin=False)
        db.add(user)
        count += 1
    db.commit()
    return {"imported": count, "skipped": skipped}


@app.post("/api/admin/testers/{tester_id}/reset")
def admin_reset_tester(tester_id: int, admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    tester = db.query(User).filter(User.id == tester_id, User.is_admin == False).first()
    if not tester:
        raise HTTPException(404)
    sessions = db.query(TestSession).filter(TestSession.user_id == tester.id, TestSession.status != "voided").all()
    for s in sessions:
        s.status = "voided"
    db.commit()
    return {"ok": True}


@app.delete("/api/admin/testers/{tester_id}")
def admin_delete_tester(tester_id: int, admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    tester = db.query(User).filter(User.id == tester_id, User.is_admin == False).first()
    if not tester:
        raise HTTPException(404)
    db.delete(tester)
    db.commit()
    return {"ok": True}


# ─── Admin: results ────────────────────────────────────

@app.get("/api/admin/results")
def admin_list_results(
    language_pair_id: Optional[int] = Query(None),
    tester_id: Optional[str] = Query(None),
    admin: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    q = db.query(TestSession).filter(TestSession.status != "voided")
    if language_pair_id:
        q = q.filter(TestSession.language_pair_id == language_pair_id)
    if tester_id:
        user = db.query(User).filter(User.user_id == tester_id).first()
        if user:
            q = q.filter(TestSession.user_id == user.id)

    sessions = q.all()
    result = []
    for s in sessions:
        user = db.query(User).filter(User.id == s.user_id).first()
        lp = db.query(LanguagePair).filter(LanguagePair.id == s.language_pair_id).first()
        ratings = db.query(SentenceRating).filter(SentenceRating.test_session_id == s.id).all()
        # 按引擎计算总分：引擎1=自研，引擎2=讯飞
        # 如果左边是引擎1，则 user_rating < 0 表示引擎1胜（取反得正）
        # 如果右边是引擎1，则 user_rating > 0 表示引擎1胜
        engine1_score = 0
        for r in ratings:
            if r.user_rating is None:
                continue
            if r.engine_left == 'self_research':
                engine1_score -= r.user_rating  # 取反
            else:
                engine1_score += r.user_rating
        avg_score = engine1_score / len(ratings) if ratings else 0

        result.append({
            "session_id": s.id,
            "user_id": user.user_id if user else "",
            "language_pair": lp.display_name if lp else "",
            "pair_code": lp.pair_code if lp else "",
            "started_at": s.started_at.isoformat() if s.started_at else None,
            "completed_at": s.completed_at.isoformat() if s.completed_at else None,
            "total_score": engine1_score,
            "avg_score": round(avg_score, 2),
            "rated_count": len(ratings),
            "status": s.status,
        })
    return result


@app.get("/api/admin/results/{session_id}/detail")
def admin_result_detail(session_id: int, admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    session = db.query(TestSession).filter(TestSession.id == session_id).first()
    if not session:
        raise HTTPException(404)

    ratings = (
        db.query(SentenceRating)
        .filter(SentenceRating.test_session_id == session_id)
        .order_by(SentenceRating.sentence_index)
        .all()
    )
    user = db.query(User).filter(User.id == session.user_id).first()
    lp = db.query(LanguagePair).filter(LanguagePair.id == session.language_pair_id).first()

    details = []
    for r in ratings:
        sentence = db.query(Sentence).filter(Sentence.id == r.sentence_id).first()
        details.append({
            "sentence_index": r.sentence_index,
            "sid": sentence.sid if sentence else "",
            "source_text": sentence.source_text if sentence else "",
            "engine_left": r.engine_left,
            "engine_right": r.engine_right,
            "left_play_count": r.left_play_count,
            "right_play_count": r.right_play_count,
            "user_rating": r.user_rating,
            "duration_seconds": r.duration_seconds,
            "rated_at": r.rated_at.isoformat() if r.rated_at else None,
        })

    return {
        "user_id": user.user_id if user else "",
        "language_pair": lp.display_name if lp else "",
        "pair_code": lp.pair_code if lp else "",
        "status": session.status,
        "started_at": session.started_at.isoformat() if session.started_at else None,
        "completed_at": session.completed_at.isoformat() if session.completed_at else None,
        "details": details,
    }


# ─── Export Excel ───────────────────────────────────────

@app.get("/api/export/excel")
def export_excel(
    session_id: Optional[int] = Query(None),
    language_pair_id: Optional[int] = Query(None),
    tester_id: Optional[str] = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    wb = Workbook()

    # Sheet 1: detail
    ws1 = wb.active
    ws1.title = "单句明细"
    headers1 = [
        "test_user_id", "language_pair", "corpus_id", "source_text",
        "source_audio_duration", "engine_left", "engine_right",
        "left_translation_text", "right_translation_text",
        "left_recognition_text", "right_recognition_text",
        "left_audio_play_count", "right_audio_play_count",
        "user_rating", "sentence_duration_seconds", "timestamp",
    ]
    ws1.append(headers1)

    q = db.query(TestSession).filter(TestSession.status != "voided")
    if session_id:
        q = q.filter(TestSession.id == session_id)
    if language_pair_id:
        q = q.filter(TestSession.language_pair_id == language_pair_id)
    if tester_id:
        u = db.query(User).filter(User.user_id == tester_id).first()
        if u:
            q = q.filter(TestSession.user_id == u.id)
    if not user.is_admin:
        q = q.filter(TestSession.user_id == user.id)

    sessions = q.all()
    summary_rows = []

    for sess in sessions:
        u = db.query(User).filter(User.id == sess.user_id).first()
        lp = db.query(LanguagePair).filter(LanguagePair.id == sess.language_pair_id).first()
        ratings = (
            db.query(SentenceRating)
            .filter(SentenceRating.test_session_id == sess.id)
            .order_by(SentenceRating.sentence_index)
            .all()
        )

        engine1_score = 0
        engine1_better2 = 0  # 自研更好
        engine1_better1 = 0  # 自研好一点
        ties = 0
        engine2_better1 = 0  # 讯飞好一点
        engine2_better2 = 0  # 讯飞更好

        for r in ratings:
            sentence = db.query(Sentence).filter(Sentence.id == r.sentence_id).first()

            if r.engine_left == "self_research":
                lt = sentence.engine1_translation_text if sentence else ""
                lr = sentence.engine1_recognition_text if sentence else ""
                rt = sentence.engine2_translation_text if sentence else ""
                rr = sentence.engine2_recognition_text if sentence else ""
            else:
                lt = sentence.engine2_translation_text if sentence else ""
                lr = sentence.engine2_recognition_text if sentence else ""
                rt = sentence.engine1_translation_text if sentence else ""
                rr = sentence.engine1_recognition_text if sentence else ""

            ws1.append([
                u.user_id if u else "",
                lp.pair_code if lp else "",
                sentence.sid if sentence else "",
                sentence.source_text if sentence else "",
                sentence.source_audio_duration if sentence else 0,
                r.engine_left,
                r.engine_right,
                lt, rt, lr, rr,
                r.left_play_count,
                r.right_play_count,
                r.user_rating,
                r.duration_seconds,
                r.rated_at.strftime("%Y-%m-%d %H:%M:%S") if r.rated_at else "",
            ])

            # 按引擎统计
            if r.user_rating is not None:
                if r.engine_left == 'self_research':
                    # 左边是自研，rating < 0 表示自研胜
                    engine1_score -= r.user_rating
                    if r.user_rating == -2:
                        engine1_better2 += 1
                    elif r.user_rating == -1:
                        engine1_better1 += 1
                    elif r.user_rating == 0:
                        ties += 1
                    elif r.user_rating == 1:
                        engine2_better1 += 1
                    elif r.user_rating == 2:
                        engine2_better2 += 1
                else:
                    # 右边是自研，rating > 0 表示自研胜
                    engine1_score += r.user_rating
                    if r.user_rating == -2:
                        engine2_better2 += 1
                    elif r.user_rating == -1:
                        engine2_better1 += 1
                    elif r.user_rating == 0:
                        ties += 1
                    elif r.user_rating == 1:
                        engine1_better1 += 1
                    elif r.user_rating == 2:
                        engine1_better2 += 1

        avg = engine1_score / len(ratings) if ratings else 0
        first_time = ratings[0].rated_at if ratings else None
        last_time = ratings[-1].rated_at if ratings else None
        duration_min = (last_time - first_time).total_seconds() / 60 if first_time and last_time else 0

        summary_rows.append([
            u.user_id if u else "",
            lp.pair_code if lp else "",
            first_time.strftime("%Y-%m-%d %H:%M:%S") if first_time else "",
            last_time.strftime("%Y-%m-%d %H:%M:%S") if last_time else "",
            round(duration_min, 1),
            engine1_score,
            round(avg, 2),
            engine1_better2, engine1_better1, ties, engine2_better1, engine2_better2,
            sess.status,
        ])

    # Sheet 2: summary
    ws2 = wb.create_sheet("汇总统计")
    headers2 = [
        "test_user_id", "language_pair", "test_start_time", "test_end_time",
        "total_duration_minutes", "engine1_score(自研)", "avg_score",
        "自研更好", "自研好一点", "差不多", "讯飞好一点", "讯飞更好", "status",
    ]
    ws2.append(headers2)
    for row in summary_rows:
        ws2.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"evaluation_results_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── Serve frontend static files ───────────────────────

FRONTEND_DIST = os.path.join(os.path.dirname(__file__), "..", "frontend", "dist")
if os.path.isdir(FRONTEND_DIST):
    app.mount("/assets", StaticFiles(directory=os.path.join(FRONTEND_DIST, "assets")), name="static-assets")

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        file_path = os.path.join(FRONTEND_DIST, full_path)
        if full_path and os.path.isfile(file_path):
            return FileResponse(file_path)
        return FileResponse(os.path.join(FRONTEND_DIST, "index.html"))
